import openpyxl


def createTestNum():
    d = ["1", "2", "3", "4", "5", "6", "7"]
    return d
 
def writeExcel(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()  
    sheet = workbook.active  
    sheet.title = sheet_name  
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
    workbook.save(path)  
    print("xlsx格式表格写入数据成功！")
 
def saveExcel():
    try:
        data = openpyxl.load_workbook('测试数据.xlsx')  
        sheetnames = data.get_sheet_names() # 取第一张表
        table = data.get_sheet_by_name(sheetnames[0]) 
        table = data.active
        print(table.title)  # 输出表名

        nrows = table.max_row  # 获得行数
        # ncolumns = table.max_column  # 获得列数
    except:
        writeExcel(book_name_xlsx,sheet_name_xlsx,value3)

    numList = createTestNum() # 获取测试数据
    print(numList)
    values = [
    numList
    ]

    # 注意行业列下标是从1开始的
    for i in range(1, len(values)+1):
        for j in range(1, len(values[i-1])+1):
            table.cell(nrows+i, j).value = values[i-1][j-1]
    
    data.save('测试数据.xlsx')


if __name__ == "__main__":
    book_name_xlsx = '测试数据.xlsx'
    sheet_name_xlsx = '测试数据表'
    value3 = [["姓名", "性别", "年龄", "城市", "职业"],
          ["111", "女", "66", "石家庄", "运维工程师"],
          ["222", "男", "55", "南京", "饭店老板"],
          ["333", "女", "27", "苏州", "保安"], ]
    for i in range(10):
        # writeExcel(book_name_xlsx, sheet_name_xlsx, value3)
        saveExcel()